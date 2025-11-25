"""
Парсер BTR-файлов в формате Excel.
Поддерживает три различных формата файлов с рейтингами BTR.
Поддерживает как новый формат .xlsx, так и старый .xls.
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Union

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import xlrd

logger = logging.getLogger(__name__)


# Маппинг названий листов на категории рейтинга
SHEET_MAPPINGS = {
    # Формат 1: короткие английские названия
    "M": "men_double",
    "Mmx": "men_mixed",
    "W": "women_double",
    "Wmx": "women_mixed",
    "YM19": "junior_male",
    "YW19": "junior_female",
    # Формат 2: полные русские названия
    "Взрослые, парный, мужчины": "men_double",
    "Взрослые, смешанный, мужчины": "men_mixed",
    "Взрослые, парный, женщины": "women_double",
    "Взрослые, смешанный, женщины": "women_mixed",
    "До 19, Юноши": "junior_male",
    "До 19, Девушки": "junior_female",
    # Формат 3: короткие русские названия
    "М": "men_double",
    "ММикст": "men_mixed",
    "Ж": "women_double",
    "ЖМикст": "women_mixed",
    "Ю 19": "junior_male",
    "Д 19": "junior_female",
}


class BtrPlayerData:
    """Данные игрока из BTR-файла."""

    def __init__(
        self,
        category: str,
        rank: int,
        rating_value: float,
        last_name: str,
        first_name: str,
        middle_name: str,
        rni: int,
        birth_date: Optional[datetime],
        city: str,
        tournaments_total: int,
        tournaments_52_weeks: int,
        tournaments_counted: int,
    ):
        self.category = category
        self.rank = rank
        self.rating_value = rating_value
        self.last_name = last_name
        self.first_name = first_name
        self.middle_name = middle_name
        self.rni = rni
        self.birth_date = birth_date
        self.city = city
        self.tournaments_total = tournaments_total
        self.tournaments_52_weeks = tournaments_52_weeks
        self.tournaments_counted = tournaments_counted
        # Автоматически определяем пол по категории
        self.gender = _determine_gender_from_category(category)

    def __repr__(self):
        return f"<BtrPlayerData {self.last_name} {self.first_name} ({self.category}, РНИ: {self.rni}): {self.rating_value}>"


def _parse_date(value) -> Optional[datetime]:
    """Парсит дату из различных форматов."""
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, str):
        # Попробуем различные форматы
        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue

    return None


def _parse_number(value, default=0) -> int:
    """Парсит число из различных форматов."""
    if value is None or value == "":
        return default

    if isinstance(value, (int, float)):
        return int(value)

    if isinstance(value, str):
        try:
            # Убираем пробелы и заменяем запятую на точку
            cleaned = value.strip().replace(" ", "").replace(",", ".")
            return int(float(cleaned))
        except ValueError:
            return default

    return default


def _parse_rating(value) -> float:
    """Парсит рейтинг (может быть дробным числом)."""
    if value is None or value == "":
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        try:
            # Убираем пробелы и заменяем запятую на точку
            cleaned = value.strip().replace(" ", "").replace(",", ".")
            return float(cleaned)
        except ValueError:
            return 0.0

    return 0.0


def _determine_gender_from_category(category: str) -> str:
    """
    Определяет пол игрока по категории рейтинга.
    
    Args:
        category: Категория рейтинга (men_double, women_double, etc.)
        
    Returns:
        'male' или 'female'
    """
    if category in ("men_double", "men_mixed", "junior_male"):
        return "male"
    elif category in ("women_double", "women_mixed", "junior_female"):
        return "female"
    else:
        return "male"  # По умолчанию


class SheetAdapter:
    """Универсальный адаптер для работы с листами openpyxl и xlrd."""
    
    def __init__(self, sheet, is_xls=False):
        self.sheet = sheet
        self.is_xls = is_xls
        if is_xls:
            self.max_row = sheet.nrows
        else:
            self.max_row = sheet.max_row
    
    def get_row(self, row_num):
        """Получить строку (1-indexed для совместимости с openpyxl)."""
        if self.is_xls:
            # xlrd использует 0-indexed
            if row_num - 1 >= self.sheet.nrows:
                return []
            return self.sheet.row_values(row_num - 1)
        else:
            # openpyxl использует 1-indexed
            return [cell.value for cell in self.sheet[row_num]]
    
    def get_cell_value(self, row_num, col_num):
        """Получить значение ячейки (1-indexed)."""
        if self.is_xls:
            # xlrd использует 0-indexed
            if row_num - 1 >= self.sheet.nrows or col_num - 1 >= self.sheet.ncols:
                return None
            cell = self.sheet.cell(row_num - 1, col_num - 1)
            # Обработка дат в xlrd
            if cell.ctype == 3:  # XL_CELL_DATE
                try:
                    date_tuple = xlrd.xldate_as_tuple(cell.value, 0)
                    return datetime(*date_tuple)
                except:
                    return cell.value
            return cell.value
        else:
            # openpyxl использует 1-indexed
            return self.sheet.cell(row_num, col_num).value


def _detect_format_and_header_row(sheet: Union[Worksheet, SheetAdapter]) -> Tuple[int, Optional[Dict[str, int]]]:
    """
    Определяет формат файла и находит строку с заголовками.
    Возвращает (номер_строки_с_заголовками, маппинг_колонок).
    """
    # Адаптируем sheet если это не SheetAdapter
    if not isinstance(sheet, SheetAdapter):
        sheet = SheetAdapter(sheet, is_xls=False)
    
    # Формат 1 и 2: заголовки в первой строке
    first_row = sheet.get_row(1)
    if "Место" in first_row and "Очки" in first_row and "Фамилия" in first_row:
        # Формат 1 или 2
        col_map = {}
        for idx, val in enumerate(first_row, start=1):
            if val == "Место":
                col_map["rank"] = idx
            elif val == "Очки" and "rating" not in col_map:  # Берём первый столбец "Очки"
                col_map["rating"] = idx
            elif val == "Фамилия":
                col_map["last_name"] = idx
            elif val == "Имя":
                col_map["first_name"] = idx
            elif val == "Отчество":
                col_map["middle_name"] = idx
            elif val == "РНИ":
                col_map["rni"] = idx
            elif val == "ДеньРождения":
                col_map["birth_date"] = idx
            elif val == "Город":
                col_map["city"] = idx
            elif val == "ВсегоСыграно":
                col_map["tournaments_total"] = idx
            elif val == "РазрядСыграно":
                col_map["tournaments_52_weeks"] = idx
            elif val == "Учтено":
                col_map["tournaments_counted"] = idx
        return 1, col_map

    # Формат 3: заголовки в 10-й или 11-й строке
    for row_num in [10, 11]:
        if row_num > sheet.max_row:
            continue
        row = sheet.get_row(row_num)
        if "№" in row and "Очки" in row and ("ФИО" in row or "Ф" in row):
            col_map = {}
            for idx, val in enumerate(row, start=1):
                if val == "№":
                    col_map["rank"] = idx
                elif val == "Очки" and "rating" not in col_map:  # Берём первый столбец "Очки"
                    col_map["rating"] = idx
                elif val == "ФИО":
                    col_map["fio"] = idx
                elif val == "Ф":
                    col_map["last_name"] = idx
                elif val == "И":
                    col_map["first_name"] = idx
                elif val == "О":
                    col_map["middle_name"] = idx
                elif val == "РНИ":
                    col_map["rni"] = idx
                elif val == "Дата рождения":
                    col_map["birth_date"] = idx
                elif val == "Город":
                    col_map["city"] = idx
                elif val and "Кол-во сыгранных турниров" in str(val):
                    col_map["tournaments_total"] = idx
                elif val and "Кол-во турниров за 52 нед" in str(val):
                    col_map["tournaments_52_weeks"] = idx
                elif val and "Кол-во учтенных турниров" in str(val):
                    col_map["tournaments_counted"] = idx
            return row_num, col_map

    return 0, None


def _parse_fio(fio_str: str) -> Tuple[str, str, str]:
    """
    Парсит ФИО из одной строки.
    Возвращает (фамилия, имя, отчество).
    """
    if not fio_str:
        return "", "", ""

    parts = fio_str.strip().split()
    last_name = parts[0] if len(parts) > 0 else ""
    first_name = parts[1] if len(parts) > 1 else ""
    middle_name = parts[2] if len(parts) > 2 else ""
    return last_name, first_name, middle_name


def _parse_sheet(sheet: Union[Worksheet, SheetAdapter], category: str) -> List[BtrPlayerData]:
    """Парсит один лист Excel-файла."""
    players = []

    # Адаптируем sheet если это не SheetAdapter
    if not isinstance(sheet, SheetAdapter):
        sheet = SheetAdapter(sheet, is_xls=False)

    # Определяем формат и находим заголовки
    header_row, col_map = _detect_format_and_header_row(sheet)
    if not col_map:
        logger.warning(f"Не удалось определить формат листа '{sheet.title}'")
        return players

    # Читаем данные, начиная со строки после заголовков
    for row_num in range(header_row + 1, sheet.max_row + 1):
        # Проверяем, что строка не пустая
        rank_val = sheet.get_cell_value(row_num, col_map.get("rank", 1))
        if rank_val is None or rank_val == "":
            continue

        # Парсим данные
        rank = _parse_number(rank_val)
        rating_value = _parse_rating(sheet.get_cell_value(row_num, col_map.get("rating", 2)))

        # ФИО может быть в одном столбце или в трёх разных
        if "fio" in col_map:
            fio_str = sheet.get_cell_value(row_num, col_map["fio"]) or ""
            last_name, first_name, middle_name = _parse_fio(str(fio_str))
        else:
            last_name = str(sheet.get_cell_value(row_num, col_map.get("last_name", 3)) or "").strip()
            first_name = str(sheet.get_cell_value(row_num, col_map.get("first_name", 4)) or "").strip()
            middle_name = str(sheet.get_cell_value(row_num, col_map.get("middle_name", 5)) or "").strip()

        rni = _parse_number(sheet.get_cell_value(row_num, col_map.get("rni", 6)))
        birth_date = _parse_date(sheet.get_cell_value(row_num, col_map.get("birth_date", 7)))
        city = str(sheet.get_cell_value(row_num, col_map.get("city", 8)) or "").strip()
        tournaments_total = _parse_number(sheet.get_cell_value(row_num, col_map.get("tournaments_total", 10)))
        tournaments_52_weeks = _parse_number(sheet.get_cell_value(row_num, col_map.get("tournaments_52_weeks", 11)))
        tournaments_counted = _parse_number(sheet.get_cell_value(row_num, col_map.get("tournaments_counted", 12)))

        # Пропускаем строки без РНИ или фамилии
        if not rni or rni == 0 or not last_name:
            continue

        player = BtrPlayerData(
            category=category,
            rank=rank,
            rating_value=rating_value,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
            rni=rni,
            birth_date=birth_date,
            city=city,
            tournaments_total=tournaments_total,
            tournaments_52_weeks=tournaments_52_weeks,
            tournaments_counted=tournaments_counted,
        )
        players.append(player)

    return players


def parse_btr_file(file_path: str, rating_date: datetime) -> List[BtrPlayerData]:
    """
    Парсит BTR-файл и возвращает список данных игроков.
    Поддерживает как .xlsx (openpyxl), так и .xls (xlrd) форматы.

    Args:
        file_path: Путь к Excel-файлу
        rating_date: Дата рейтинга

    Returns:
        Список объектов BtrPlayerData
    """
    file_path_obj = Path(file_path)
    if not file_path_obj.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    logger.info(f"Парсинг файла: {file_path}")

    # Определяем формат файла по расширению или содержимому
    is_xls = False
    workbook = None
    
    try:
        # Сначала пробуем открыть как .xlsx
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
    except Exception as xlsx_error:
        # Если не получилось, пробуем как .xls
        logger.info(f"Не удалось открыть как .xlsx, пробуем .xls: {xlsx_error}")
        try:
            workbook = xlrd.open_workbook(file_path, formatting_info=False)
            sheet_names = workbook.sheet_names()
            is_xls = True
        except Exception as xls_error:
            logger.error(f"Не удалось открыть файл ни как .xlsx, ни как .xls: {xls_error}")
            raise ValueError(f"Неподдерживаемый формат файла: {file_path}")

    all_players = []
    
    # Все ожидаемые категории
    expected_categories = {
        "men_double",
        "men_mixed",
        "women_double",
        "women_mixed",
        "junior_male",
        "junior_female",
    }
    found_categories = set()

    # Обрабатываем только известные листы
    for sheet_name in sheet_names:
        # Убираем пробелы в начале и конце названия листа
        sheet_name_clean = sheet_name.strip()
        if sheet_name_clean in SHEET_MAPPINGS:
            category = SHEET_MAPPINGS[sheet_name_clean]
            logger.info(f"Обработка листа '{sheet_name}' -> категория '{category}'")
            
            if is_xls:
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_adapter = SheetAdapter(sheet, is_xls=True)
            else:
                sheet = workbook[sheet_name]
                sheet_adapter = SheetAdapter(sheet, is_xls=False)
            
            players = _parse_sheet(sheet_adapter, category)
            all_players.extend(players)
            found_categories.add(category)
            logger.info(f"Найдено {len(players)} игроков в категории '{category}'")

    # Проверяем, все ли категории найдены
    missing_categories = expected_categories - found_categories
    if missing_categories:
        logger.warning(
            f"ВНИМАНИЕ! В файле {file_path_obj.name} не найдены следующие категории: "
            f"{', '.join(sorted(missing_categories))}. "
            f"Доступные листы: {', '.join(sheet_names[:10])}"
        )

    if not is_xls:
        workbook.close()

    logger.info(f"Всего найдено {len(all_players)} записей игроков из {len(found_categories)}/6 категорий")
    return all_players
